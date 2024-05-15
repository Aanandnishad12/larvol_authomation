import re
import datetime
def strip_str(res):
    msg = '##############################   extra space at start and end    ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            temp_data = data.strip()
            
            if not data == temp_data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'
def Encoding_str(res):
    msg = '##############################   Encoding string    ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            temp_data = data
            
            if '_x00' in temp_data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'

def mid_format(res):
    msg = '##############################   invalid mid format   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            try:
                temp_data = re.search('(([A-Za-z]+_)+\d+)',data).group(1)
                
                if not data == temp_data:
                    msg += f'{key} ============>  Bad format   {temp_data} ================>Row no. {ct}\n'
            except:
                msg += f'{key} ============>  Bad format   {data} ================>Row no. {ct}\n'

            
    return f'{msg}\n\n'
def line_breakes(res):
    msg = '##############################   line_breakes    ##################################\n\n'
    
    for key in res:
        ct = 1
        for data in res[key]:
            ct += 1
            
            temp_data = re.sub('\r','',data,flags=re.S)
            temp_data = re.sub('\n','',data,flags=re.S)
            
            if not data == temp_data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'

def author_aff_delimiter(res):
    msg = '##############################   Invalid author_aff_delimiter   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            temp_data = re.sub('\s*;\s*','; ',data,flags=re.S)
            
            if not data == temp_data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'  
def space_(res):
    msg = '##############################   space in mid time url   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if ' 'in data:
                msg += f'{key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'   
def date_format(res):
    msg = '##############################   Invalid date_format   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if not data:
                continue
            try:
                datetime.datetime.strptime(data, '%B %d, %Y')
                
            except:
                msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'
            
            else:
                # October 18, 2022
                if not re.search('([A-Z][a-z]+ \d\d, \d\d\d\d)',data):
                    msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'

            
    return f'{msg}\n\n'      

def start_end_time_format(res):
    msg = '##############################   Invalid start_end_time_format   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            
            if not data:
                continue
            try:
                datetime.datetime.strptime(data, '%H:%M')
                
            except:
                msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'
                
            else:
                if not re.search('(\d\d:\d\d)',data):
                    msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'
        
    return f'{msg}\n\n'  

def start_end_time_les_0_6(res):
    msg = '##############################   time ( 7 =< time >= 0 )   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if not data:
                continue
            try:
                temp_data = data.replace(' ','').replace(':','').strip()
                if int(temp_data) <= 700:
                    msg += f'{key} ============>    this early time not possible {data}   ===========>   Row no. {ct}\n'
            except:
                continue
                        
    return f'{msg}\n\n' 
def is_paid_wrong_data(res):
    msg = '############################   Wrong data in is_paid   ##############################\n\n'

    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if data=='No' or data=="Yes":
                continue

            else:
                msg += f'{key} ============>    Wrong data in is_paid {data}   ===========>   Row no. {ct}\n'

        return f'{msg}\n\n'
def start_end_time_morethan_2330(res):
    msg = '##############################   time ( time >= 23:30 )   ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            if not data:
                continue
            try:

                temp_data = data.replace(' ','').replace(':','').strip()
                if int(temp_data) >= 2330:
                    msg += f'{key} ============>    this late time not possible {data}   ===========>   Row no. {ct}\n'
            except:
                continue
                        
    return f'{msg}\n\n' 
def if_end_time_small(res):
    msg = '##############################   Start time is grater then End time  ##################################\n\n'

    ct = 1
    for x in range(len(res['end_time'])):
        ct += 1
        if res['end_time'][x] == '':
            continue
        
        temp_end = int(''.join(re.findall('\d',res['end_time'][x],flags=re.S)))
        temp_start = int(''.join(re.findall('\d',res['start_time'][x],flags=re.S)))

        if temp_start > temp_end :
                msg += f'end_time ============>   start_time end_time is same    ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n' 
def invalid_end_time(res):
    msg = '##############################   start_time is blank but end_time is there   ##################################\n\n'

    ct = 1
    for x in range(len(res['start_time'])):
        ct += 1
        if not res['start_time'][x]:
            if res['end_time'][x]:
                msg += f'end_time ============>   end time not possible    ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n'  
def start_time_end_time(res):
    msg = '##############################   start_time end_time is same   ##################################\n\n'

    ct = 1
    for x in range(len(res['start_time'])):
        ct += 1
        if res['start_time'][x]== '':
            continue

        if res['start_time'][x] == res['end_time'][x]:
                msg += f'end_time ============>   start_time end_time is same    ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n'    
def session_id_format(res):
    msg = '##############################   invalid session id format   ##################################\n\n'

    ct = 1
    for x in range(len(res['session_id'])):
        ct += 1
        if res['session_id'][x]== '':
            continue
        
        if not re.search('^S\d+$',res['session_id'][x],flags=re.S):
                msg += f'session_id ============>   invalid session id format {res["session_id"][x]}  ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n'  
def news_type_format(res):
    msg = '##############################   invalid news_type format   ##################################\n\n'

    ct = 1
    for x in range(len(res['news_type'])):
        ct += 1
        if res['news_type'][x]== '':
            continue
        
        if not re.search('^Session$|^Abstract$',res['news_type'][x],flags=re.S):
                msg += f'news_type ============>   invalid news_type format {res["news_type"][x]}  ===========>   Row no. {ct}\n'
                
                        
    return f'{msg}\n\n'
def Formula_in_excel(res):
    msg = '##############################   Formula in excel (AS General)  ##################################\n\n'

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    df = pd.read_excel(res)

    workbook = load_workbook(res)
    worksheet = workbook.active

    for column in df.columns:
        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
        ct = 0
        for cell in worksheet[column_letter]:
            ct += 1
            if cell.data_type == 'f':
                msg += f"Formula found in column '{column}': {cell.value} row number {ct}\n"
        ct = 0
                        
    return f'{msg}\n\n'     
def strip_str(res):
    msg = '##############################   Formula in excel (AS Text)    ##################################\n\n'
    
    for key in res:
        
        ct = 1
        for data in res[key]:
            ct += 1
            
            temp_data = re.search('(^=[A-Z]+\()',data,flags=re.S)
            if temp_data:
                temp_data = re.search('(^=[A-Z]+\(.*)',data,flags=re.S)
                msg += f'Formula found in column \'{key}\': {temp_data.group(1)} row number {ct}\n'
            
            
    return f'{msg}\n\n'

def Two_or_more_session_as_new_in_sid(df):
    msg = '##############################   Two or more Session(news_type) for single SID   ##################################\n\n'

    grouped_bk_group = df.groupby('session_id')

    for _, group_df in grouped_bk_group:
        if group_df['news_type'].to_list().count('Session')>1:

            msg += f' For SID \"{_}\" there is two or more Session in news_type coln for single SID please check  \n\n'

    return f'{msg}\n\n'
def multiple_session_title_to_single_sid(df):
    msg = '##############################   multiple session_title to single SID   ##################################\n\n'

    grouped_bk_group = df.groupby('session_id')

    for _, group_df in grouped_bk_group:
        if len(set(group_df['session_title'].to_list()))>1:

            msg += f' SID \"{_}\" is pointing to multiple different session titles please check  \n\n'

    return f'{msg}\n\n'
sponsor = ["Sponsor",
"Sponsored by",
"Sponsorship",
"Fund",
"Funded by",
"Financed",
"Financed by",
"Financial support",
"Supported by",
"Acknowledgement",
"Acknowledged by",
"Registration ID",
"Clinical Trial ID"]


sponsor_msg = ''


def sponsor_col(res):
    msg = '##############################   sponsor   ##################################\n\n'
    
    for key in res:
        ct = 1
        for data in res[key]:
            ct += 1
            
            for sp_key in sponsor:
                if sp_key.lower() in data.lower():

                    msg += f'May can be sponsor {sp_key} ==========> {key} ============> Row no. {ct}\n'
            
            
    return f'{msg}\n\n'