from flask import Flask, render_template, request, redirect, url_for, send_file, flash
from werkzeug.utils import secure_filename
import pandas as pd
import numpy as np


import re
#from Config import *
from datetime import datetime
import datetime
# import os
import sys
import os
from io import BytesIO
from zipfile import ZipFile
# from flask import Flask, render_template, request, send_file
# import pandas as pd
from openpyxl import load_workbook, Workbook
# from io import BytesIO
from excelclear import excel_clear
from seperateauthor import amanin
app = Flask(__name__)

app.config['SECRET_KEY'] = os.environ.get('FLASK_SECRET_KEY')
ALLOWED_EXTENSIONS = {'xlsx'}

UPLOAD_FOLDERS =  "uploads2"
app.config['UPLOAD_FOLDERS2'] = UPLOAD_FOLDERS
UPLOAD_FOLDERS =  "uploads1"
app.config['UPLOAD_FOLDERS1'] = UPLOAD_FOLDERS
UPLOAD_FOLDERS = 'uploads3'
app.config['UPLOAD_FOLDER3'] = UPLOAD_FOLDERS
MAX_FILE_SIZE = 15* 1024 * 1024  # 12 MB


def clear_uploads_folder(UPLOAD_FOLDER):
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    for filename in os.listdir(UPLOAD_FOLDER):
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        try:
            if os.path.isfile(file_path) or os.path.islink(file_path):
                os.unlink(file_path)
            elif os.path.isdir(file_path):
                os.rmdir(file_path)
        except Exception as e:
            print(f'Failed to delete {file_path}. Reason: {e}')



def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/', methods=['GET', 'POST'])
def hello_world():
    if request.method == 'POST':
        
        if 'file' not in request.files:
            return 'No file part'

        file = request.files['file']
        number = request.form["choice"]

        if file.filename == '':
            return 'No selected file'
        file.seek(0, os.SEEK_END)
        file_length = file.tell()
        

        if file_length > MAX_FILE_SIZE:
            return "error File size exceeds the limit of 15 MB"

        file.seek(0)
                

        try:
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # df = pd.DataFrame(file)
                # print(df)

                save_directory = os.path.dirname(os.path.abspath(__file__))
                save_directory+="/Previus"
                if os.path.exists(save_directory) and os.path.isdir(save_directory):
                    # Get a list of all the files and subdirectories in the directory
                    for item in os.listdir(save_directory):
                        item_path = os.path.join(save_directory, item)
                        
                        if os.path.isfile(item_path):
                            # If it's a file, remove it
                            os.remove(item_path)
                        elif os.path.isdir(item_path):
                            # If it's a directory, recursively empty it
                            save_directory(item_path)
                
                # print(save_directory,"0000000000000000000000000000000000000000000000000000000000000")
                file.save(os.path.join(save_directory, filename))
                print(save_directory)
                source_file = fr"{save_directory}/{filename}"
                print(source_file)
                
                df = pd.read_excel(source_file)
                df = df.replace(np.nan, '')
                # uploaded_df = pd.read_excel(source_file, sheet_name=None, header=1)
                uploaded_df = pd.read_excel(source_file, sheet_name=None, header=1)
                # uploaded_df = uploaded_df.replace(np.nan, '')
                # df = pd.read_excel(source_file)
                # df = df.replace(np.nan, '')
                # print(df)
                source_file = source_file.split('.')[0]

                if os.path.exists(f'{source_file}.txt'):
                    os.remove(f'{source_file}.txt')
                    print(f'deleting priviouslg generated comments {source_file}.txt')

                to_write = ''

                if not df.columns.tolist() == ['is_paid','source_id', 'manual_id', 'article_title', 'url', 'authors', 'author_affiliation', 'abstract_text', 'date', 'start_time', 'end_time', 'location', 'session_id', 'news_type',
                'session_title', 'session_type', 'category', 'sub_category', 'disclosure']:
                    print('hi there')
                    return 'hi there row header is wrong'
                    to_write += '#########################################################################\n\n row header is wrong \n\n#########################################################################'
                    sys.exit("row header is wrong")



                df = df.astype(str)
                is_paid = df['is_paid'].tolist()
                source_id = df['source_id'].tolist()
                manual_id = df['manual_id'].tolist()
                url = df['url'].tolist()
                article_title = df['article_title'].tolist()
                authors = df['authors'].tolist()
                author_affiliation = df['author_affiliation'].tolist()
                abstract_text = df['abstract_text'].tolist()
                date = df['date'].tolist()
                start_time = df['start_time'].tolist()
                end_time = df['end_time'].tolist()
                location = df['location'].tolist()
                session_title = df['session_title'].tolist()
                session_type = df['session_type'].tolist()
                category = df['category'].tolist()
                sub_category = df['sub_category'].tolist()
                disclosure = df['disclosure'].tolist()
                session_id = df['session_id'].tolist()
                news_type = df['news_type'].tolist()


                res = {'source_id':source_id,
                'manual_id': manual_id,
                'url':url,
                'article_title':article_title,
                'authors':authors,
                'author_affiliation':author_affiliation,
                'abstract_text':abstract_text,
                'date':date,
                'start_time':start_time,
                'end_time':end_time,
                'location':location,
                'session_title':session_title,
                'session_id':session_id,
                'news_type':news_type,
                'session_type':session_type,
                'category':category,
                'sub_category':sub_category,
                'disclosure':disclosure}

                paitron_msg = ''

                    
                def strip_str(res):
                    msg = '##############################   extra space at start and end    ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            temp_data = data.strip()
                            
                            if not data == temp_data:
                                msg += f'{key} ============> Row no. {ct}\n'
                            
                            
                    return f'{msg}\n\n'

                temp_msg = strip_str(res)
                paitron_msg += temp_msg
                to_write += temp_msg


                def Encoding_str(res):
                    msg = '##############################   Encoding string    ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            temp_data = data
                            
                            if '_x00' in temp_data:
                                msg += f'{key} ============> Row no. {ct}\n'
                            
                            
                    return f'{msg}\n\n'

                temp_msg = Encoding_str(res)
                paitron_msg += temp_msg
                to_write += temp_msg



                def mid_format(res):
                    msg = '##############################   invalid mid format   ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            try:
                                temp_data = re.search('(([A-Za-z]+_)+\d+)',data).group(1)
                                
                                if not data == temp_data:
                                    msg += f'{key} ============>  Bad format   {temp_data} ================>Row no. {ct}\n'
                            except:
                                msg += f'{key} ============>  Bad format   {data} ================>Row no. {ct}\n'

                            
                    return f'{msg}\n\n'

                temp_msg = mid_format({'manual_id': manual_id})
                paitron_msg += temp_msg
                to_write += temp_msg



                def line_breakes(res):
                    msg = '##############################   line_breakes    ##################################\n\n'
                    
                    for key in res:
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            
                            temp_data = re.sub('\r','',data,flags=re.S)
                            temp_data = re.sub('\n','',data,flags=re.S)
                            
                            if not data == temp_data:
                                msg += f'{key} ============> Row no. {ct}\n'
                            
                            
                    return f'{msg}\n\n'

                temp_msg = line_breakes(res)
                paitron_msg += temp_msg
                to_write += temp_msg


                def author_aff_delimiter(res):
                    msg = '##############################   Invalid author_aff_delimiter   ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            temp_data = re.sub('\s*;\s*','; ',data,flags=re.S)
                            
                            if not data == temp_data:
                                msg += f'{key} ============> Row no. {ct}\n'
                            
                            
                    return f'{msg}\n\n'  

                temp_msg = (author_aff_delimiter({'authors':authors, 'author_affiliation':author_affiliation}))
                paitron_msg += temp_msg
                to_write += temp_msg


                def space_(res):
                    msg = '##############################   space in mid time url   ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            if ' 'in data:
                                msg += f'{key} ============> Row no. {ct}\n'
                            
                            
                    return f'{msg}\n\n'   

                temp_msg = space_({'manual_id':manual_id,'url':url, 'start_time':start_time, 'end_time':end_time})
                paitron_msg += temp_msg
                to_write += temp_msg



                def date_format(res):
                    msg = '##############################   Invalid date_format   ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            if not data:
                                continue
                            try:
                                datetime.datetime.strptime(data, '%B %d, %Y')
                                
                            except:
                                msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'
                            
                            else:
                                # October 18, 2022
                                if not re.search('([A-Z][a-z]+ \d\d, \d\d\d\d)',data):
                                    msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'

                            
                    return f'{msg}\n\n'      

                temp_msg = (date_format({'date':date}))
                paitron_msg += temp_msg
                to_write += temp_msg
                def current_year_and_date_check(res):
                    msg = '##############################   year check! Correct All date format before year checking ##################################\n\n'
                    # print(list(set(res["date"])),"----------------------")

                    date_format = "%B %d, %Y"
                    ct = 1
                    for date_str in res["date"]:
                        ct+=1
                        # Parse the string into a datetime object
                        try:
                            date_obj = datetime.datetime.strptime(date_str, date_format)
                            # Get the current year
                            current_year = datetime.datetime.now().year
                            
                            # Check if the year matches the current year
                            if date_obj.year == current_year:
                                # return date_obj
                                pass
                            else:
                                msg += f'date ============> Row no. {ct} has {date_obj.year} rather than current year {str(current_year)}\n'
                        except:
                            pass

                    return msg

                temp_msg = (current_year_and_date_check({'date':date}))
                paitron_msg += temp_msg
                to_write += temp_msg

                def start_end_time_format(res):
                    msg = '##############################   Invalid start_end_time_format   ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            
                            if not data:
                                continue
                            try:
                                datetime.datetime.strptime(data, '%H:%M')
                                
                            except:
                                msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'
                                
                            else:
                                if not re.search('(\d\d:\d\d)',data):
                                    msg += f'{key} ============>Bad Format {data}   ===========>   Row no. {ct}\n'
                        
                    return f'{msg}\n\n'      

                temp_msg = (start_end_time_format({'start_time':start_time,'end_time':end_time}))
                paitron_msg += temp_msg
                to_write += temp_msg


                def start_end_time_les_0_6(res):
                    msg = '##############################   time ( 7 =< time >= 0 )   ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            if not data:
                                continue
                            try:
                                temp_data = data.replace(' ','').replace(':','').strip()
                                if int(temp_data) <= 700:
                                    msg += f'{key} ============>    this early time not possible {data}   ===========>   Row no. {ct}\n'
                            except:
                                continue
                                        
                    return f'{msg}\n\n'      

                temp_msg = (start_end_time_les_0_6({'start_time':start_time,'end_time':end_time}))
                paitron_msg += temp_msg
                to_write += temp_msg

                def is_paid_wrong_data(res):
                    msg = '############################   Wrong data in is_paid   ##############################\n\n'

                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            if data=='No' or data=="Yes":
                                continue

                            else:
                                msg += f'{key} ============>    Wrong data in is_paid {data}   ===========>   Row no. {ct}\n'

                        return f'{msg}\n\n'
                    

                temp_msg = is_paid_wrong_data({'is_pad':is_paid})
                paitron_msg += temp_msg
                to_write += temp_msg

                def start_end_time_morethan_2330(res):
                    msg = '##############################   time ( time >= 23:30 )   ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            if not data:
                                continue
                            try:

                                temp_data = data.replace(' ','').replace(':','').strip()
                                if int(temp_data) >= 2330:
                                    msg += f'{key} ============>    this late time not possible {data}   ===========>   Row no. {ct}\n'
                            except:
                                continue
                                        
                    return f'{msg}\n\n'      

                temp_msg = (start_end_time_morethan_2330({'start_time':start_time,'end_time':end_time}))
                paitron_msg += temp_msg
                to_write += temp_msg


                def if_end_time_small(res):
                    msg = '##############################   Start time is grater then End time  ##################################\n\n'

                    ct = 1
                    for x in range(len(res['end_time'])):
                        ct += 1
                        if res['end_time'][x] == '':
                            continue
                        
                        temp_end = int(''.join(re.findall('\d',res['end_time'][x],flags=re.S)))
                        temp_start = int(''.join(re.findall('\d',res['start_time'][x],flags=re.S)))

                        if temp_start > temp_end :
                                msg += f'end_time ============>   start_time end_time is same    ===========>   Row no. {ct}\n'
                                
                                        
                    return f'{msg}\n\n' 

                temp_msg = (if_end_time_small({'start_time':start_time,'end_time':end_time}))
                paitron_msg += temp_msg
                to_write += temp_msg


                def invalid_end_time(res):
                    msg = '##############################   start_time is blank but end_time is there   ##################################\n\n'

                    ct = 1
                    for x in range(len(res['start_time'])):
                        ct += 1
                        if not res['start_time'][x]:
                            if res['end_time'][x]:
                                msg += f'end_time ============>   end time not possible    ===========>   Row no. {ct}\n'
                                
                                        
                    return f'{msg}\n\n'      

                temp_msg = (invalid_end_time({'start_time':start_time,'end_time':end_time}))
                paitron_msg += temp_msg
                to_write += temp_msg


                def start_time_end_time(res):
                    msg = '##############################   start_time end_time is same   ##################################\n\n'

                    ct = 1
                    for x in range(len(res['start_time'])):
                        ct += 1
                        if res['start_time'][x]== '':
                            continue

                        if res['start_time'][x] == res['end_time'][x]:
                                msg += f'end_time ============>   start_time end_time is same    ===========>   Row no. {ct}\n'
                                
                                        
                    return f'{msg}\n\n'      

                temp_msg = (start_time_end_time({'start_time':start_time,'end_time':end_time}))
                paitron_msg += temp_msg
                to_write += temp_msg


                def session_id_format(res):
                    msg = '##############################   invalid session id format   ##################################\n\n'

                    ct = 1
                    for x in range(len(res['session_id'])):
                        ct += 1
                        if res['session_id'][x]== '':
                            continue
                        
                        if not re.search('^S\d+$',res['session_id'][x],flags=re.S):
                                msg += f'session_id ============>   invalid session id format {res["session_id"][x]}  ===========>   Row no. {ct}\n'
                                
                                        
                    return f'{msg}\n\n'      

                temp_msg = (session_id_format({'session_id':session_id}))
                paitron_msg += temp_msg
                to_write += temp_msg

                def news_type_format(res):
                    msg = '##############################   invalid news_type format   ##################################\n\n'

                    ct = 1
                    for x in range(len(res['news_type'])):
                        ct += 1
                        if res['news_type'][x]== '':
                            continue
                        
                        if not re.search('^Session$|^Abstract$',res['news_type'][x],flags=re.S):
                                msg += f'news_type ============>   invalid news_type format {res["news_type"][x]}  ===========>   Row no. {ct}\n'
                                
                                        
                    return f'{msg}\n\n'      

                temp_msg = (news_type_format({'news_type':news_type}))
                paitron_msg += temp_msg
                to_write += temp_msg


                def Formula_in_excel(res):
                    msg = '##############################   Formula in excel (AS General)  ##################################\n\n'

                    import pandas as pd
                    from openpyxl import load_workbook
                    from openpyxl.utils import get_column_letter

                    df = pd.read_excel(res)

                    workbook = load_workbook(res)
                    worksheet = workbook.active

                    for column in df.columns:
                        column_letter = get_column_letter(df.columns.get_loc(column) + 1)
                        ct = 0
                        for cell in worksheet[column_letter]:
                            ct += 1
                            if cell.data_type == 'f':
                                msg += f"Formula found in column '{column}': {cell.value} row number {ct}\n"
                        ct = 0
                                        
                    return f'{msg}\n\n'      

                temp_msg = Formula_in_excel(source_file + '.xlsx')
                paitron_msg += temp_msg
                to_write += temp_msg


                def strip_str(res):
                    msg = '##############################   Formula in excel (AS Text)    ##################################\n\n'
                    
                    for key in res:
                        
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            
                            temp_data = re.search('(^=[A-Z]+\()',data,flags=re.S)
                            if temp_data:
                                temp_data = re.search('(^=[A-Z]+\(.*)',data,flags=re.S)
                                msg += f'Formula found in column \'{key}\': {temp_data.group(1)} row number {ct}\n'
                            
                            
                    return f'{msg}\n\n'

                temp_msg = strip_str(res)
                paitron_msg += temp_msg
                to_write += temp_msg


                def Two_or_more_session_as_new_in_sid(df):
                    msg = '##############################   Two or more Session(news_type) for single SID   ##################################\n\n'

                    grouped_bk_group = df.groupby('session_id')

                    for _, group_df in grouped_bk_group:
                        if group_df['news_type'].to_list().count('Session')>1:

                            msg += f' For SID \"{_}\" there is two or more Session in news_type coln for single SID please check  \n\n'

                    return f'{msg}\n\n'

                temp_msg = Two_or_more_session_as_new_in_sid(df)
                paitron_msg += temp_msg
                to_write += temp_msg


                def multiple_session_title_to_single_sid(df):
                    msg = '##############################   multiple session_title to single SID   ##################################\n\n'

                    grouped_bk_group = df.groupby('session_id')

                    for _, group_df in grouped_bk_group:
                        if len(set(group_df['session_title'].to_list()))>1:

                            msg += f' SID \"{_}\" is pointing to multiple different session titles please check  \n\n'

                    return f'{msg}\n\n'


                temp_msg = multiple_session_title_to_single_sid(df)
                paitron_msg += temp_msg
                to_write += temp_msg



                if '' in manual_id:
                    temp_msg = '=================>  vacant cell found in manual_id column  <=================\n\n'
                    to_write += temp_msg
                    
                    
                if '' in url:
                    temp_msg = '=================>  vacant cell found in url column   <=================\n\n'
                    to_write += temp_msg
                    
                if '' in article_title:
                    temp_msg = '=================>  vacant cell found in article_title column   <=================\n\n'
                    to_write += temp_msg

                if '' in is_paid:
                    temp_msg = '=================>  vacant cell found in is_paid column   <=================\n\n'
                    to_write += temp_msg

                manual_id_to_b_unique = df["source_id"].tolist()
                source_id_to_b_unique = df["manual_id"].tolist()
                article_title_to_b_unique = df["article_title"].tolist()

                manual_id_unique_dic = {}
                source_id_unique_dic = {}
                article_title_unique_dic = {}

                for x in range(len(manual_id_to_b_unique)):
                    if not manual_id_unique_dic.get(manual_id_to_b_unique[x],''):
                        manual_id_unique_dic[manual_id_to_b_unique[x]] = 1
                    else: 
                        manual_id_unique_dic[manual_id_to_b_unique[x]] += 1

                    if not source_id_unique_dic.get(source_id_to_b_unique[x],''):
                        source_id_unique_dic[source_id_to_b_unique[x]] = 1
                    else: 
                        source_id_unique_dic[source_id_to_b_unique[x]] += 1
                        
                    if not article_title_unique_dic.get(article_title_to_b_unique[x],''):
                        article_title_unique_dic[article_title_to_b_unique[x]] = 1
                    else: 
                        article_title_unique_dic[article_title_to_b_unique[x]] += 1
                        
                        
                to_write_unique_manual_id = ''
                to_write_unique_source_id = ''
                to_write_unique_article_title = ''


                for x,y in manual_id_unique_dic.items():
                    if y>1 and x!='':
                        to_write_unique_manual_id = to_write_unique_manual_id + f'{x} ===========> total count is {y}\n'
                        
                        
                for x,y in source_id_unique_dic.items():
                    if y>1 and x!='':
                        to_write_unique_source_id = to_write_unique_source_id + f'{x} ===========> total count is {y}\n'
                        
                for x,y in article_title_unique_dic.items():
                    if y>1 and x!='':
                        to_write_unique_article_title = to_write_unique_article_title + f'{x} ===========> total count is {y}\n'

                to_write_2 = f'''##############################   Duplicate mid  ##################################\n\n
                        mannual_id which are not unique \n{to_write_unique_source_id}\n\n\n\n\n\n
                -------------------------------------------------------------------------------------------------------------------------------------------------------------------
                #################################################################        Duplicates          ######################################################################
                -------------------------------------------------------------------------------------------------------------------------------------------------------------------
                        source_id which are not unique \n{to_write_unique_manual_id}\n\n\n\n\n\n
                        
                        artical_id which are not unique \n{to_write_unique_article_title}'''

                to_write_2 = re.sub('1qaz2wsx.*?total count is \d+','',to_write_2)

                to_write += '\n' + to_write_2


                with open(f"{source_file}_QC_comments.txt",'w',encoding='utf-8') as f:
                    f.write(to_write)

                sponsor = ["Sponsor",
                "Sponsored by",
                "Sponsorship",
                "Fund",
                "Funded by",
                "Financed",
                "Financed by",
                "Financial support",
                "Supported by",
                "Acknowledgement",
                "Acknowledged by",
                "Registration ID",
                "Clinical Trial ID"]


                sponsor_msg = ''


                def sponsor_col(res):
                    msg = '##############################   sponsor   ##################################\n\n'
                    
                    for key in res:
                        ct = 1
                        for data in res[key]:
                            ct += 1
                            
                            for sp_key in sponsor:
                                if sp_key.lower() in data.lower():

                                    msg += f'May can be sponsor {sp_key} ==========> {key} ============> Row no. {ct}\n'
                            
                            
                    return f'{msg}\n\n'

                temp_msg = sponsor_col({"abstract_text":abstract_text})
                sponsor_msg += temp_msg

                with open(f'{source_file}_sponsor.txt','w',encoding='utf-8')as s_dtails:
                    s_dtails.write(sponsor_msg)


                # Process or save the uploaded file here
                xlsx_file_path = fr"{source_file}_QC_comments.txt"  # Update this with the actual file path
     
                xlsx_file_path = fr"{source_file}_QC_comments.txt"  # Update this with the actual file path

                xlsx_file_path = fr"{source_file}_QC_comments.txt"
                with open(xlsx_file_path, "r") as file1:
                    content1 = file1.read()

                # Read the second text file
                xlsx_file_path = fr"{source_file}_sponsor.txt"
                with open(xlsx_file_path, "r") as file2:
                    content2 = file2.read()

                # Combine the contents of both files
                context = f"{content1}\n\n{content2}" if content1 and content2 else (content1 or content2)


                return render_template("index.html", context=context)
        except Exception as e:
            print(e,"=============================================================")
            return e


    return render_template("index.html")



@app.route('/excel_cleaner', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':

        if 'file' not in request.files:
            return "nice"

        file = request.files['file']
        print(file.filename,"++++++++++++++++++++++++")
        file.seek(0, os.SEEK_END)
        file_length = file.tell()
        
        text_input = request.form.get('text_input', '').strip()
        if file_length > MAX_FILE_SIZE:
            return "error File size exceeds the limit of 15 MB"
        if len(text_input)==0:
            text_input = "All Speakers"

        file.seek(0)
        

        if file and allowed_file(file.filename):
            clear_uploads_folder(app.config['UPLOAD_FOLDERS1'])
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDERS1'], filename))
            # file.save(os.path.join(app.config['UPLOAD_FOLDERS2'], filename))
            print(os.path.join(app.config['UPLOAD_FOLDERS1']),"===========================")
            namessddsd = str(file.filename).split(".")[0]
            excel_clear(namessddsd,group_by_author_replace=text_input)
            return render_template('excel.html', download_ready=True)
            # return send_file(fr"./uploads1/{namessddsd}_colored.xlsx", as_attachment=True, download_name=fr"./uploads1/{namessddsd}_colored.xlsx")
            # return render_template('excel.html')
    return render_template('excel.html')

@app.route('/download')
def download_file():
    try:
        # Extract query parameters
        # namessddsd = request.args.get('namessddsd')
        directory = './uploads1'

# Get all files in the directory
        files = os.listdir(directory)
        for i in files:
            if "_colored" in i:
                print(fr"./uploads1/{i}")
                return send_file(fr"./uploads1/{i}", as_attachment=True, download_name=fr"./uploads1/{i}")
        else:
            return "hahaha i m joking"
    except Exception as e:
        return "error Error sending the file " 
    

@app.route('/separate_authors', methods=['GET', 'POST'])
def upload_file1():
    if request.method == 'POST':

        if 'file' not in request.files:
            return "nice"

        file = request.files['file']
        print(file.filename,"++++++++++++++++++++++++")
        file.seek(0, os.SEEK_END)
        file_length = file.tell()
        
        text_input = request.form.get('text_input', '').strip()
        if file_length > MAX_FILE_SIZE:
            return "error File size exceeds the limit of 15 MB"

        file.seek(0)
        

        if file and allowed_file(file.filename):
            clear_uploads_folder(app.config['UPLOAD_FOLDERS2'])
            filename = secure_filename(file.filename)
            file.save(os.path.join(app.config['UPLOAD_FOLDERS2'], filename))
            # file.save(os.path.join(app.config['UPLOAD_FOLDERS2'], filename))
            print(os.path.join(app.config['UPLOAD_FOLDERS2']),"===========================")
            # namessddsd = str(file.filename).split(".")[0]
            # excel_clear(namessddsd,group_by_author_replace=text_input)
            dfdff = pd.read_excel(fr"./uploads2/{file.filename}")
            dfdff = dfdff.replace(np.nan, '')
            # dfdff.dropna(subset=['authors'], inplace=True)

            # Filter rows where 'authors' column contains '\d+'
            filtered_df = dfdff[dfdff['authors'].str.contains(r'\d+')]
            print(filtered_df)
            remaining_df = dfdff[~dfdff.index.isin(filtered_df.index)]
            print(remaining_df)
            filtered_df.to_excel("./uploads2/1.xlsx")
            remaining_df.to_excel("./uploads2/2.xlsx")
            amanin("./uploads2/1.xlsx","./uploads2/1next.xlsx",1)
            amanin("./uploads2/2.xlsx","./uploads2/2next.xlsx",2)
            df11 = pd.read_excel("./uploads2/1next.xlsx")
            df22 = pd.read_excel("./uploads2/2next.xlsx")
            df_final = pd.concat([df11, df22])
            df_final['numeric_id'] = df_final['manual_id'].str.extract(r'(\d+)').astype(int)

# Sort DataFrame by 'numeric_id' and reset index
            df_sorted = df_final.sort_values(by='numeric_id').reset_index(drop=True)

            # Remove the temporary 'numeric_id' column
            df_sorted.drop(columns='numeric_id', inplace=True)

            # print(df_sorted)
            df_sorted = df_sorted.applymap(lambda x: x.strip() if isinstance(x, str) else x)
            df_sorted.to_excel("./uploads2/finaauthor.xlsx")




            
            return render_template('sepauthor.html', download_ready=True)
    return render_template('sepauthor.html')
@app.route('/download1')
def download_file1():
    # return "hello"
    try:


        return send_file(fr"./uploads2/finaauthor.xlsx", as_attachment=True, download_name=fr"./uploads2/finaauthor.xlsx")

    except Exception as e:
        return "error Error sending the file " 
    

def upload_file():
    if 'files[]' not in request.files:
        flash('No file part')
        return redirect(request.url)
    
    files = request.files.getlist('files[]')
    for file in files:
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        else:
            flash('Allowed file types are .xlsx')
            return redirect(request.url)
    
    flash('Files successfully uploaded')
    return redirect(url_for('index'))

if __name__ == "__main__":
    app.run(host='0.0.0.0')

