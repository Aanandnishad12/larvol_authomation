#!/usr/bin/env python
# coding: utf-8
import warnings

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module='bs4')
from bs4 import BeautifulSoup

import os
import pandas as pd
import html
import re
import dateutil.parser
from datetime import datetime
from configparser import ConfigParser
import ftfy
import openpyxl
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import PatternFill
    



def strip_( cell):

    # cell = BeautifulSoup(str(cell), "lxml").text

    cell = html.unescape(cell)
    cell = cell.encode("utf_16")
    cell = cell.decode("utf_16")
    cell = ftfy.fix_text(cell)
    # cell = re.sub('<.*?>','',str(cell),flags=re.S)
    cell = re.sub('\s*;\s*','; ',str(cell),flags=re.S)
    cell = re.sub('\s+',' ',str(cell),flags=re.S)
    cell = cell.replace('\n',' ')
    cell = cell.replace('\r','')
    cell = re.sub('\s+',' ',str(cell),flags=re.S)
    cell = cell.strip()

    if cell == ' ' or cell == '':
        return ''
    return cell

def trim_dataframe(df):
    df = df.astype(str)
    df = df.applymap(strip_)
    df = df.applymap(strip_)
    return df

def convert_df(name):
    import numpy as np
	
    df = pd.read_excel(fr'./uploads1/{name}.xlsx',engine='openpyxl')
    df.reset_index(drop=True, inplace=True)
    df = df.replace(np.nan, '')
    return df

def clean_excel(df):

    df = trim_dataframe(df)
    return df

def df_excel(df,name):

    df.to_excel(f'./uploads1/{name}_polished.xlsx',index=False,engine='openpyxl')

def date_time_formetor_manuall (data,manual_date_format,desire_format):
    date_time_obj = datetime.strptime(data, manual_date_format)
    # ist_date_time = date_time_obj - timedelta(hours = 0,minutes = 0)  
    ist_date_time = ist_date_time.strftime(desire_format)
    return ist_date_time

def date_time_formetor(data,desire_format):

    try:
        yourdate = dateutil.parser.parse(data)
        oldformat = yourdate
        datetimeobject = datetime.strptime(str(oldformat),'%Y-%m-%d  %H:%M:%S')
        newformat = datetimeobject.strftime(desire_format)
    #    print(data,'==================',newformat)
        return newformat
    except:
        return f'Bad Format {data}'

def remove_str_nam(df):
    import numpy as np

    # df = df.applymap(lambda x: '1qaz2wsx' if str(x) == 'nan' else x ) #x.replace('nan','1qaz2wsx'))
    df = df.replace(np.nan, '')
    # df.fillna('NA',inplace=True)
    return df


def excel_clear(target,sheet_name="Acronym_2024_LT_MMDDYYYY",group_by="session_title",group_by_author_replace="All Speakers",author_affiliation_special_end_removal=";",manual_date_format=False):
    # target =Acronym_2024_LT_MMDDYYYY_colored_MID_colored
    # sheet_name = Acronym_2024_LT_MMDDYYYY

    # group by col for removing garbage author



    # author_affiliation_special_end_removal can be false or special symbol that should be removed at the end
    author_affiliation_special_end_removal = author_affiliation_special_end_removal

    manual_date_format = manual_date_format

    #Get the password
    # details = config_object["details"]
    excel_name = target
    # allow_unique_author = details["allow_unique_author"]
    manual_date_format = manual_date_format
    author_affiliation_special_end_removal = author_affiliation_special_end_removal
    sheet_name = sheet_name
    group_by = group_by
    group_by_author_replace = group_by_author_replace


    print('selected sheet ===> ',sheet_name)
    print('mannual date format ===> ', manual_date_format)
    if not manual_date_format == 'false':
        date_flag = True
    else:
        date_flag = False

    print('excel created with name ===> ',excel_name)
    print('sheet name is ===> ', sheet_name)
    print('grouping by ===> ',group_by)
    print('pre-populate author ===> ', group_by_author_replace)
    print('remove special text from author cells ===> ', author_affiliation_special_end_removal)

    df = convert_df(excel_name)

    df.columns = ["is_paid","source_id", "manual_id", "article_title", "url", "authors", "author_affiliation", "abstract_text",
                "date", "start_time", "end_time", "location", "session_id", "news_type", "session_title", "session_type", "category", "sub_category",
                "disclosure"]
    df = clean_excel(df)

    start_time = df['start_time'].tolist()
    end_time = df['end_time'].tolist()

    new_start_time = []
    new_end_time = []

    for x in range(len(start_time)):

        start_time[x] = str(start_time[x])
        if start_time[x] == 'nan':
            new_start_time.append('nan')
        else: 
            new_start_time.append(date_time_formetor(start_time[x],'%H:%M'))

        end_time[x] = str(end_time[x])
        if end_time[x] == 'nan':
            new_end_time.append('nan')
        else:       
            new_end_time.append(date_time_formetor(end_time[x],'%H:%M'))

    df['new_start_time'] = new_start_time
    df['new_end_time'] = new_end_time 

    # authors = df['authors'].tolist()
    # authors_affs = df['author_affiliation'].tolist()
    # new_author = []
    # new_aff = []

    # if allow_unique_author == 'true' :
    #     for x in range(len(authors_affs)):
    #         author = str(authors[x])
    #         authors_aff = str(authors_affs[x])

    #         author = author.split('; ')
    #         authors_aff = authors_aff.split('; ')

    #         if len(author) > 1:
    #             author = '; '.join(list(set(author)))
    #         else:
    #             author = author[0]
    #         new_author.append(author)

    #         if len(authors_aff) > 1:
    #             authors_aff = '; '.join(list(set(authors_aff)))
    #         else:
    #             authors_aff = authors_aff[0]
    #         new_aff.append(authors_aff)

    #     df['new_author'] =  new_author
    #     df['new_aff'] =  new_aff
    # else:       
    #     df['new_author'] =  ['']*len(authors_affs)
    #     df['new_aff'] =  ['']*len(authors_affs)

    date_ = df['date'].tolist()
    new_date = []

    for x in range(len(date_)):
        date_[x] = str(date_[x])

        if date_[x] == 'nan':
            new_date.append('')
        else: 
            if date_flag :
                try:
                    new_date.append(date_time_formetor_manuall(date_[x],manual_date_format,'%B %d, %Y'))
                except:
                    new_date.append(date_time_formetor(date_[x],'%B %d, %Y'))
            else :
                new_date.append(date_time_formetor(date_[x],'%B %d, %Y'))


    df['new_date'] = new_date    

    # df = df[["source_id", "manual_id", "article_title", "url", "authors","new_author", "author_affiliation","new_aff", "abstract_text",
    #          "date", 'new_date', "start_time","new_start_time", "end_time","new_end_time", "location", "session_title", "session_type", "category", "sub_category",
    #          "disclosure"]]

    authors = df['authors'].tolist()
    author_affiliation = df['author_affiliation'].tolist()
    if author_affiliation_special_end_removal != 'false':
        author_affiliation_special_end_removal = author_affiliation_special_end_removal.replace('-','\-').replace('+','\+').replace('.','\.').replace('*','\*').replace('?','\?').replace('[','\[').replace(']','\]')
        for x in range(len(authors)):
            authors[x] = re.sub(f'\s*{author_affiliation_special_end_removal}\s*$','',authors[x],flags = re.S)
            author_affiliation[x] = re.sub(f'{author_affiliation_special_end_removal}\s*$','',author_affiliation[x],flags = re.S)

    df['authors'] = authors
    df['author_affiliation'] = author_affiliation

    def chk_aff_already_added(to_find_aut, to_find_aff, aut_list, aff_list)-> bool:
        '''
        True :- aff and aut already exist
        False :- aff not added
        '''

        for aut, aff in zip(aut_list, aff_list):
            if aut == to_find_aut and aff == to_find_aff:
                return True

        return False



    # filling all authors
    if group_by_author_replace.lower() != 'false':

    #     group_by = "session_title"#enter the name of column to group by
    #     group_by_author_replace = "All Participants"#eg all authors

        
    #     if not (group_by in df):
    #         print(f'{group_by} is not available in excel file')
    #         print(f'try using one from {list(df.columns)}')

        group_by_lst = group_by.split('=>')

            
        grp = df.groupby(group_by_lst)


        for g in grp:
            if str(g[0]) and str(g[0]).lower() != 'nan':#avoiding empty values and nan values
                g_df = g[1].copy()
                all_authors = g_df['authors'].astype(str).to_list()
                all_aff = g_df['author_affiliation'].astype(str).to_list()


                final_aut = []
                final_aff = []

                for aut,aff in zip(all_authors, all_aff):
                    for k,v in zip(aut.split('; '), aff.split('; ')):
                        if k != 'nan' and k != group_by_author_replace:#if aut is valid
                            if v == 'nan':
                                if k not in final_aut:
                                    final_aut.append(k)
                                    final_aff.append('')
                            else:
                                if not chk_aff_already_added(k, v, final_aut, final_aff):
                                    final_aut.append(k)
                                    final_aff.append(v)
        #         uncoment to change
    #             print(all_authors)
    #             print(all_aff)
    #             print('after processing---------------')
    #             print('; '.join(list(_dict.keys())))
    #             print('; '.join(list(_dict.values())))
    #             print('\n____________________')

                temp_df = g[1][g[1]['authors'] == group_by_author_replace]
                
                df.loc[temp_df.index, ['author_affiliation']] = '; '.join(final_aff)

                df.loc[temp_df.index, ['authors']] = '; '.join(final_aut)



    df = df[["is_paid","source_id", "manual_id", "article_title", "url", "authors", "author_affiliation", "abstract_text",
                "date","new_date", "start_time", "new_start_time", "end_time","new_end_time", "location", "session_id","news_type", "session_title", "session_type",
            "category", "sub_category", "disclosure"]]

    df = remove_str_nam(df)

    df_excel(df,excel_name)

    try:
        wb = load_workbook(f'./uploads1/{excel_name}_polished.xlsx', data_only=True)
    except:
        print('Not available to read')

    try:
        default_sheet = wb['Sheet1']#try
    except:
        print('Sheet not present try choosing one from:',wb.sheetnames)


    from openpyxl.styles.borders import Border, Side, BORDER_THIN
    thin_border = thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='00000000'),
        right=Side(border_style=BORDER_THIN, color='00000000'),
        top=Side(border_style=BORDER_THIN, color='00000000'),
        bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

    ct = 0
    color_list = ['B8CCE4','DBE5F1']
    for row in default_sheet.rows:

        if ct == 0:
            ct+= 2
            for cell in row: #4F81BD
                cell.fill = PatternFill(start_color="4F81BD", \
                                                end_color="4F81BD", fill_type = 'solid')
                cell.border = thin_border
            continue
        for cell in row:
            cell.fill = PatternFill(start_color=color_list[ct%2], \
                                                end_color=color_list[ct%2], fill_type = 'solid')
            cell.border = thin_border
        ct+=1

    try:
        wb.save(f'./uploads1/{excel_name}_colored.xlsx')
        os.remove(f"./uploads1/{excel_name}_polished.xlsx")
        os.remove(f"./uploads1/{excel_name}_unique_mid_sid_title.txt")
    except:
        print('Not available to read')

        


    manual_id_to_b_unique = df["source_id"].tolist()
    source_id_to_b_unique = df["manual_id"].tolist()
    article_title_to_b_unique = df["article_title"].tolist()

    manual_id_unique_dic = {}
    source_id_unique_dic = {}
    article_title_unique_dic = {}

    for x in range(len(manual_id_to_b_unique)):
        if not manual_id_unique_dic.get(manual_id_to_b_unique[x],''):
            manual_id_unique_dic[manual_id_to_b_unique[x]] = 1
        else: 
            manual_id_unique_dic[manual_id_to_b_unique[x]] += 1

        if not source_id_unique_dic.get(source_id_to_b_unique[x],''):
            source_id_unique_dic[source_id_to_b_unique[x]] = 1
        else: 
            source_id_unique_dic[source_id_to_b_unique[x]] += 1
            
        if not article_title_unique_dic.get(article_title_to_b_unique[x],''):
            article_title_unique_dic[article_title_to_b_unique[x]] = 1
        else: 
            article_title_unique_dic[article_title_to_b_unique[x]] += 1
            
            
    to_write_unique_manual_id = ''
    to_write_unique_source_id = ''
    to_write_unique_article_title = ''


    for x,y in manual_id_unique_dic.items():
        if y>1:
            to_write_unique_manual_id = to_write_unique_manual_id + f'{x} ===========> total count is {y}\n'
            
            
    for x,y in source_id_unique_dic.items():
        if y>1:
            to_write_unique_source_id = to_write_unique_source_id + f'{x} ===========> total count is {y}\n'
            
    for x,y in article_title_unique_dic.items():
        if y>1:
            to_write_unique_article_title = to_write_unique_article_title + f'{x} ===========> total count is {y}\n'

    to_write = f'''           source_id which are not unique \n{to_write_unique_manual_id}\n\n\n\n\n\n
            mannual_id which are not unique \n{to_write_unique_source_id}\n\n\n\n\n\n
            artical_id which are not unique \n{to_write_unique_article_title}'''

    to_write = re.sub('1qaz2wsx.*?total count is \d+','',to_write)

    to_write = str(to_write)
    with open(f"./uploads1/{excel_name}_unique_mid_sid_title.txt",'w',encoding='utf-8') as f:
        f.write(to_write)
        
    print('=='*30+'\n')      
    print('process ended\n')


